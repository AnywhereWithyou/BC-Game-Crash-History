# Import required modules
from datetime import datetime
import os
import time
from playwright.sync_api import Playwright, sync_playwright
import openpyxl


def CrashHistory(url, rowNo, rowDelta):
    print(url, str(rowNo), str(rowDelta))
    # Set up Playwright options
    options = {
        'headless': False,  # Run in headless mode to avoid opening a new window
    }
    # Launch a new browser instance
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(**options)
        page = browser.new_page()

        # Navigate to the website with a longer timeout value
        page.goto(url)
        time.sleep(3)

        # Wait for the page to finish loading
        page.wait_for_load_state('networkidle')

        # Find the input with id="game_amount_input" and enter the number of rows to retrieve
        input_element = page.locator('#game_amount_input')
        input_element.fill(str(rowDelta))

        # Find the button with id="game_verify_submit" and click it
        button_element = page.locator('#game_verify_submit')
        button_element.click()

        time.sleep(60)

        # Wait for the page to finish loading
        page.wait_for_load_state('networkidle')

        # Find all tr elements in the tbody
        tbody_locator = page.locator("#game_verify_table")
        tr_locators = tbody_locator.locator("tr")
        tr_elements = tr_locators.all()

        # Get the current datetime and format it as a string
        # now = datetime.now()
        # dt_string = now.strftime("%Y%m%d_%H%M%S")

        # Create a folder named "BC History" if it does not exist
        folder_name = "BC History"
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        # Set the name of the Excel file based on the current datetime and rowNo
        excel_file_name = f"{rowNo}_{rowNo-rowDelta+1}.xlsx"
                # excel_file_name = f"{dt_string}_{rowNo}_{rowNo-rowDelta+1}.xlsx"

        # Set the path to the Excel file inside the "BC History" folder
        excel_file_path = os.path.join(folder_name, excel_file_name)

        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Verify Table"

        # Add the headers to the worksheet
        worksheet.cell(row=1, column=1, value="Row No")
        worksheet.cell(row=1, column=2, value="Hash Content")
        worksheet.cell(row=1, column=3, value="Bang Content")

        # Loop over the tr elements and get the contents of the first two td elements
        i = 1
        for tr_element in tr_elements:
            i = i + 1
            td_locators = tr_element.locator("td")
            td_elements = td_locators.all()
            td1_content = td_elements[0].inner_text()
            td2_content = td_elements[1].inner_text()

            # Print the row number, td1_content, and td2_content
            print(rowNo, td1_content, td2_content)

            # Add the data to the worksheet
            worksheet.cell(row=i, column=1, value=rowNo)
            worksheet.cell(row=i, column=2, value=td1_content)
            worksheet.cell(row=i, column=3, value=td2_content)

            rowNo -= 1

        # Save the workbook to a file
        workbook.save(excel_file_path)
        print("************************ The end ************************")

        # Close the browser instance
        browser.close()
